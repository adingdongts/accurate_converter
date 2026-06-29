import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from copy import copy
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="Accurate SO Converter",
    page_icon="📊",
    layout="centered"
)

st.markdown("""
<style>
    .main { max-width: 720px; margin: auto; }
    .stApp { background-color: #f8f9fb; }
    .block-container { padding-top: 2rem; padding-bottom: 2rem; }
    .info-box {
        background: #eef2ff;
        border-left: 4px solid #4f46e5;
        border-radius: 4px;
        padding: 0.75rem 1rem;
        font-size: 0.9rem;
        color: #3730a3;
        margin-bottom: 1rem;
    }
    .success-box {
        background: #ecfdf5;
        border-left: 4px solid #10b981;
        border-radius: 4px;
        padding: 0.75rem 1rem;
        font-size: 0.9rem;
        color: #065f46;
    }
    .stat-card {
        background: white;
        border: 1px solid #e5e7eb;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

# ── Header ──────────────────────────────────────────────────────────────────
st.markdown("## 📊 Accurate SO Converter")
st.markdown("Ubah data Excel transaksi menjadi format import **Sales Order Accurate** secara otomatis.")
st.divider()

# ── Step 1: Upload Template Accurate ────────────────────────────────────────
st.markdown("### 1 · Upload template Accurate")
st.markdown(
    '<div class="info-box">File template <code>.xlsx</code> dari Accurate '
    '(yang punya 3 baris HEADER / ITEM / EXPENSE di atas).</div>',
    unsafe_allow_html=True,
)
template_file = st.file_uploader(
    "Template Accurate (.xlsx)", type=["xlsx"], key="template",
    label_visibility="collapsed"
)

# ── Step 2: Upload Data Input ────────────────────────────────────────────────
st.markdown("### 2 · Upload data transaksi")
st.markdown(
    '<div class="info-box">File Excel dengan kolom: '
    '<code>Tgl · No Transaksi · ID Pelanggan · Kode Produk · Qty · Satuan · Harga Satuan</code></div>',
    unsafe_allow_html=True,
)
input_file = st.file_uploader(
    "Data input (.xlsx)", type=["xlsx"], key="input",
    label_visibility="collapsed"
)

# ── Preview ──────────────────────────────────────────────────────────────────
if input_file:
    try:
        preview_df = pd.read_excel(input_file)
        input_file.seek(0)
        preview_df.columns = [
            'Tgl', 'No Transaksi', 'ID Pelanggan',
            'Kode Produk', 'Qty', 'Satuan', 'Harga Satuan'
        ]
        st.markdown("**Preview data input:**")
        st.dataframe(preview_df.head(8), use_container_width=True, hide_index=True)

        n_orders = preview_df['No Transaksi'].nunique()
        n_items  = len(preview_df)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                f'<div class="stat-card"><div style="color:#6b7280;font-size:0.8rem">Total Order</div>'
                f'<div style="font-size:1.8rem;font-weight:700;color:#1e1b4b">{n_orders}</div></div>',
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                f'<div class="stat-card"><div style="color:#6b7280;font-size:0.8rem">Total Item</div>'
                f'<div style="font-size:1.8rem;font-weight:700;color:#1e1b4b">{n_items}</div></div>',
                unsafe_allow_html=True,
            )
        st.markdown("")
    except Exception as e:
        st.error(f"Gagal membaca file input: {e}")

# ── Convert ──────────────────────────────────────────────────────────────────
st.markdown("### 3 · Konversi")

def convert(template_bytes: bytes, input_bytes: bytes) -> bytes:
    # Load template
    template_wb = load_workbook(BytesIO(template_bytes))
    template_ws = template_wb.active

    # Load input
    df = pd.read_excel(BytesIO(input_bytes))
    df.columns = [
        'Tgl', 'No Transaksi', 'ID Pelanggan',
        'Kode Produk', 'Qty', 'Satuan', 'Harga Satuan'
    ]
    df['Tgl']           = pd.to_datetime(df['Tgl']).dt.strftime('%d/%m/%Y')
    df['Qty']           = df['Qty'].astype(int)
    df['Harga Satuan']  = df['Harga Satuan'].astype(int)
    df['No Transaksi']  = df['No Transaksi'].astype(str)
    df['ID Pelanggan']  = df['ID Pelanggan'].astype(str)
    df['Kode Produk']   = df['Kode Produk'].astype(str)

    # Output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Template Pesanan Penjualan'

    # Copy 3 baris header dari template (dengan style)
    for row_idx in range(1, 4):
        for col_idx, cell in enumerate(template_ws[row_idx], 1):
            new_cell = out_ws.cell(row=row_idx, column=col_idx)
            new_cell.value = cell.value
            if cell.font:      new_cell.font      = copy(cell.font)
            if cell.fill:      new_cell.fill      = copy(cell.fill)
            if cell.alignment: new_cell.alignment = copy(cell.alignment)

    for col_letter, col_dim in template_ws.column_dimensions.items():
        out_ws.column_dimensions[col_letter].width = col_dim.width

    header_fill = PatternFill('solid', start_color='1F7041', end_color='1F7041')
    header_font = Font(bold=True, color='FFFFFF')
    item_fill   = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    item_font   = Font(bold=True, color='FFFFFF')

    order_keys  = list(dict.fromkeys(df['No Transaksi'].tolist()))
    current_row = 4

    for no_so in order_keys:
        group = df[df['No Transaksi'] == no_so]
        first = group.iloc[0]

        # HEADER row
        for col_idx, val in {
            1: 'HEADER', 2: no_so,
            3: first['Tgl'], 4: first['ID Pelanggan']
        }.items():
            c = out_ws.cell(row=current_row, column=col_idx, value=val)
            c.fill, c.font = header_fill, header_font
        current_row += 1

        # ITEM rows (no EXPENSE)
        for _, item in group.iterrows():
            for col_idx, val in {
                1: 'ITEM',           2: item['Kode Produk'],
                4: item['Qty'],      5: item['Satuan'],
                6: item['Harga Satuan']
            }.items():
                c = out_ws.cell(row=current_row, column=col_idx, value=val)
                c.fill, c.font = item_fill, item_font
            current_row += 1

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf.read()


if template_file and input_file:
    if st.button("⚡ Konversi sekarang", type="primary", use_container_width=True):
        with st.spinner("Memproses..."):
            try:
                result_bytes = convert(template_file.read(), input_file.read())
                st.markdown(
                    '<div class="success-box">✅ Konversi berhasil! Klik tombol di bawah untuk download.</div>',
                    unsafe_allow_html=True,
                )
                st.download_button(
                    label="⬇️ Download hasil (SO_Accurate_Import.xlsx)",
                    data=result_bytes,
                    file_name="SO_Accurate_Import.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Konversi gagal: {e}")
else:
    st.button("⚡ Konversi sekarang", disabled=True, use_container_width=True)
    if not template_file:
        st.caption("⬆️ Upload dulu template Accurate dan data input.")

st.divider()
st.caption("Accurate SO Converter · Tidak ada data yang dikirim ke server — semua diproses lokal.")
